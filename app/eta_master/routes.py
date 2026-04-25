import logging
import re
from datetime import datetime
from io import BytesIO

from flask import Blueprint, flash, jsonify, redirect, render_template, request, url_for
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app import limiter
from app.eta_master.models import EtaMasterRecord
from app.models import db

logger = logging.getLogger(__name__)

eta_master_bp = Blueprint('eta_master', __name__, template_folder='templates')

HEADER_ALIASES = {
    # SNO variations
    'sno': 'sno',
    'serial number': 'sno',
    'serial no': 'sno',
    's.no': 'sno',
    'slno': 'sno',
    
    # PIN CODE variations
    'pin code': 'pin_code',
    'pincode': 'pin_code',
    'pin': 'pin_code',
    'postal code': 'pin_code',
    'zip code': 'pin_code',
    
    # PICKUP STATION variations
    'pickup station': 'pickup_station',
    'pick up station': 'pickup_station',
    'pickup stn': 'pickup_station',
    'pick-up station': 'pickup_station',
    'pickupstation': 'pickup_station',
    'station': 'pickup_station',
    
    # STATE/UT variations
    'state/ut': 'state_ut',
    'state ut': 'state_ut',
    'state': 'state_ut',
    'state/union territory': 'state_ut',
    'statut': 'state_ut',
    'state-ut': 'state_ut',
    
    # CITY variations
    'city': 'city',
    'city name': 'city',
    'destination city': 'city',
    
    # PICKUP LOCATION variations
    'pickup location': 'pickup_location',
    'pick up location': 'pickup_location',
    'pick-up location': 'pickup_location',
    'pickup loc': 'pickup_location',
    'pickuplocation': 'pickup_location',
    'pick up': 'pickup_location',
    'pickup': 'pickup_location',
    'source location': 'pickup_location',
    
    # DELIVERY LOCATION variations
    'delivery location': 'delivery_location',
    'delivery loc': 'delivery_location',
    'deliverylocation': 'delivery_location',
    'delivery': 'delivery_location',
    'destination location': 'delivery_location',
    'drop location': 'delivery_location',
    
    # TAT IN DAYS variations
    'tat in days': 'tat_in_days',
    'tat': 'tat_in_days',
    'tat (days)': 'tat_in_days',
    'tat days': 'tat_in_days',
    'delivery time (days)': 'tat_in_days',
    'time to deliver': 'tat_in_days',
    'turnaround time': 'tat_in_days',
    
    # ZONE variations
    'zone': 'zone',
    'region': 'zone',
    'area': 'zone',
    'zone code': 'zone',
}

REQUIRED_FIELDS = ['pin_code', 'pickup_station', 'state_ut', 'city', 'pickup_location', 'delivery_location', 'tat_in_days', 'zone']


def _normalize_header(value):
    return re.sub(r'\s+', ' ', str(value or '').strip().lower())


def _map_headers(header_row):
    mapped = {}
    for index, header in enumerate(header_row):
        normalized = _normalize_header(header)
        mapped[index] = HEADER_ALIASES.get(normalized)
    return mapped


def _cast_sno(value):
    """Cast SNO to Integer or None."""
    if value is None or value == '' or str(value).strip() == '':
        return None
    try:
        return int(float(value))  # float first to handle Excel decimals
    except (TypeError, ValueError):
        raise ValueError(f'SNO must be an integer, got: {value}')


def _cast_pincode(value):
    """Cast PIN CODE to String (6 digits exactly)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('PIN CODE is required.')
    if not re.fullmatch(r'\d{6}', text):
        raise ValueError(f'PIN CODE must be exactly 6 digits, got: {text}')
    return text


def _cast_pickup_station(value):
    """Cast PICK UP STATION to String (required, max 255 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('PICK UP STATION is required.')
    if len(text) > 255:
        raise ValueError(f'PICK UP STATION exceeds 255 characters: {len(text)} chars')
    return text


def _cast_state_ut(value):
    """Cast STATE/UT to String (required, max 100 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('STATE/UT is required.')
    if len(text) > 100:
        raise ValueError(f'STATE/UT exceeds 100 characters: {len(text)} chars')
    return text


def _cast_city(value):
    """Cast CITY to String (required, max 100 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('CITY is required.')
    if len(text) > 100:
        raise ValueError(f'CITY exceeds 100 characters: {len(text)} chars')
    return text


def _cast_pickup_location(value):
    """Cast PICK UP LOCATION to String (required, max 255 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('PICK UP LOCATION is required.')
    if len(text) > 255:
        raise ValueError(f'PICK UP LOCATION exceeds 255 characters: {len(text)} chars')
    return text


def _cast_delivery_location(value):
    """Cast DELIVERY LOCATION to String (required, max 255 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('DELIVERY LOCATION is required.')
    if len(text) > 255:
        raise ValueError(f'DELIVERY LOCATION exceeds 255 characters: {len(text)} chars')
    return text


def _cast_tat_in_days(value):
    """Cast TAT IN DAYS to Float (required, non-negative)."""
    if value is None or value == '' or str(value).strip() == '':
        raise ValueError('TAT IN DAYS is required.')
    try:
        tat = float(value)
    except (TypeError, ValueError):
        raise ValueError(f'TAT IN DAYS must be numeric, got: {value}')
    if tat < 0:
        raise ValueError(f'TAT IN DAYS cannot be negative, got: {tat}')
    return tat


def _cast_zone(value):
    """Cast ZONE to String (required, max 50 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('ZONE is required.')
    if len(text) > 50:
        raise ValueError(f'ZONE exceeds 50 characters: {len(text)} chars')
    return text


def _build_record_payload(source):
    """Validate and normalize a record payload from a form or dict-like source."""
    sno = _cast_sno(source.get('sno'))
    pin_code = _cast_pincode(source.get('pin_code'))
    pickup_station = _cast_pickup_station(source.get('pickup_station'))
    state_ut = _cast_state_ut(source.get('state_ut'))
    city = _cast_city(source.get('city'))
    pickup_location = _cast_pickup_location(source.get('pickup_location'))
    delivery_location = _cast_delivery_location(source.get('delivery_location'))
    tat_in_days = _cast_tat_in_days(source.get('tat_in_days'))
    zone = _cast_zone(source.get('zone'))

    record_key = EtaMasterRecord.build_record_key(
        pin_code,
        pickup_station,
        state_ut,
        city,
        pickup_location,
        delivery_location,
        zone,
    )

    now = datetime.utcnow()
    return {
        'record_key': record_key,
        'sno': sno,
        'pin_code': pin_code,
        'pickup_station': pickup_station,
        'state_ut': state_ut,
        'city': city,
        'pickup_location': pickup_location,
        'delivery_location': delivery_location,
        'tat_in_days': tat_in_days,
        'zone': zone,
        'source_filename': source.get('source_filename'),
        'source_row_number': source.get('source_row_number'),
        'imported_at': now,
        'updated_at': now,
    }


def _upsert_records(rows, source_filename):
    """Bulk upsert ETA master rows with PostgreSQL ON CONFLICT handling."""
    inserted = 0
    updated = 0
    skipped = 0
    errors = []

    records = []
    for row_number, row in rows:
        try:
            record_data = _build_record_payload(row)
        except ValueError as error:
            errors.append({'row': row_number, 'error': str(error)})
            continue
        except Exception as error:
            errors.append({'row': row_number, 'error': f'Unexpected row error: {error}'})
            continue

        record_data['source_filename'] = source_filename
        record_data['source_row_number'] = row_number
        records.append(record_data)

    if not records:
        return inserted, updated, skipped, errors

    try:
        record_keys = [record['record_key'] for record in records]
        existing_keys = {
            key for (key,) in db.session.query(EtaMasterRecord.record_key)
            .filter(EtaMasterRecord.record_key.in_(record_keys))
            .all()
        }
        inserted = sum(1 for record in records if record['record_key'] not in existing_keys)
        updated = len(records) - inserted

        insert_stmt = pg_insert(EtaMasterRecord).values(records)
        upsert_stmt = insert_stmt.on_conflict_do_update(
            index_elements=['record_key'],
            set_={
                'sno': insert_stmt.excluded.sno,
                'pin_code': insert_stmt.excluded.pin_code,
                'pickup_station': insert_stmt.excluded.pickup_station,
                'state_ut': insert_stmt.excluded.state_ut,
                'city': insert_stmt.excluded.city,
                'pickup_location': insert_stmt.excluded.pickup_location,
                'delivery_location': insert_stmt.excluded.delivery_location,
                'tat_in_days': insert_stmt.excluded.tat_in_days,
                'zone': insert_stmt.excluded.zone,
                'source_filename': insert_stmt.excluded.source_filename,
                'source_row_number': insert_stmt.excluded.source_row_number,
                'updated_at': datetime.utcnow(),
            },
        )
        db.session.execute(upsert_stmt)
        db.session.commit()
        logger.info("Bulk upsert complete: %s inserted, %s updated", inserted, updated)
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error during bulk upsert: {e}")
        errors.append({'row': 'batch', 'error': f'Database error: {str(e)}'})

    return inserted, updated, skipped, errors


def _redirect_to_eta_master(mode='view', page=1, per_page=100):
    return redirect(url_for('eta_master.eta_master_upload', mode=mode, page=page, per_page=per_page))


def _load_eta_master_record(record_id):
    return db.session.get(EtaMasterRecord, record_id)


def _get_pagination_params():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 100, type=int)
    return max(page, 1), max(1, min(per_page, 500))


def _paginate_eta_master(page, per_page):
    pagination = db.paginate(
        select(EtaMasterRecord).order_by(EtaMasterRecord.id.desc()),
        page=page,
        per_page=per_page,
        error_out=False,
    )

    if pagination.total and page > pagination.pages:
        pagination = db.paginate(
            select(EtaMasterRecord).order_by(EtaMasterRecord.id.desc()),
            page=pagination.pages,
            per_page=per_page,
            error_out=False,
        )

    return pagination


def _parse_workbook(file_stream):
    workbook = load_workbook(file_stream, read_only=True, data_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise ValueError('Excel file is empty.')

    header_map = _map_headers(header_row)
    required_missing = [field for field in REQUIRED_FIELDS if field not in header_map.values()]
    if required_missing:
        raise ValueError(f'Missing required columns: {", ".join(required_missing)}')

    parsed_rows = []
    for row_number, row in enumerate(rows, start=2):
        normalized = {}
        for index, value in enumerate(row):
            field_name = header_map.get(index)
            if field_name:
                normalized[field_name] = value
        parsed_rows.append((row_number, normalized))

    return parsed_rows


@eta_master_bp.route('/eta-master', methods=['GET', 'POST'])
@limiter.limit('10 per minute', methods=['POST'])
def eta_master_upload():
    summary = None
    page, per_page = _get_pagination_params()
    edit_mode = request.args.get('mode', 'view').lower() == 'edit'
    pagination = _paginate_eta_master(page, per_page)

    if request.method == 'POST':
        upload = request.files.get('file')
        if not upload or not upload.filename:
            flash('Please choose an Excel file to import.', 'error')
            return redirect(url_for('eta_master.eta_master_upload'))

        if not upload.filename.lower().endswith('.xlsx'):
            flash('Only .xlsx Excel files are supported.', 'error')
            return redirect(url_for('eta_master.eta_master_upload'))

        try:
            parsed_rows = _parse_workbook(BytesIO(upload.read()))
            inserted, updated, skipped, errors = _upsert_records(parsed_rows, upload.filename)
            summary = {
                'inserted': inserted,
                'updated': updated,
                'skipped': skipped,
                'errors': errors,
                'total_rows': len(parsed_rows),
            }
            flash(
                f'Import complete. Inserted {inserted}, updated {updated}, errors {len(errors)}.',
                'success',
            )
            pagination = _paginate_eta_master(1, per_page)
        except Exception as error:
            logger.exception('ETA master import failed')
            flash(f'Import failed: {error}', 'error')

    start_record = 0
    end_record = 0
    if pagination.total:
        start_record = ((pagination.page - 1) * pagination.per_page) + 1
        end_record = start_record + len(pagination.items) - 1

    return render_template(
        'eta_master/index.html',
        summary=summary,
        records=pagination.items,
        pagination=pagination,
        per_page=per_page,
        start_record=start_record,
        end_record=end_record,
        edit_mode=edit_mode,
    )


@eta_master_bp.route('/eta-master/records/new', methods=['POST'])
@limiter.limit('30 per minute')
def eta_master_create_record():
    page = request.form.get('page', 1, type=int)
    per_page = request.form.get('per_page', 100, type=int)

    try:
        record_data = _build_record_payload(request.form)
        record_data['source_filename'] = 'manual entry'
        record_data['source_row_number'] = None

        duplicate = db.session.query(EtaMasterRecord).filter_by(record_key=record_data['record_key']).first()
        if duplicate:
            flash('A record with the same values already exists.', 'error')
            return _redirect_to_eta_master(mode='edit', page=page, per_page=per_page)

        db.session.add(EtaMasterRecord(**record_data))
        db.session.commit()
        flash('Record added successfully.', 'success')
    except ValueError as error:
        db.session.rollback()
        flash(f'Add failed: {error}', 'error')
    except Exception as error:
        db.session.rollback()
        logger.exception('ETA master create failed')
        flash(f'Add failed: {error}', 'error')

    return _redirect_to_eta_master(mode='view', page=1, per_page=per_page)


@eta_master_bp.route('/eta-master/records/<int:record_id>/update', methods=['POST'])
@limiter.limit('30 per minute')
def eta_master_update_record(record_id):
    page = request.form.get('page', 1, type=int)
    per_page = request.form.get('per_page', 100, type=int)

    record = _load_eta_master_record(record_id)
    if not record:
        flash('Record not found.', 'error')
        return _redirect_to_eta_master(mode='view', page=page, per_page=per_page)

    try:
        record_data = _build_record_payload(request.form)
        duplicate = db.session.query(EtaMasterRecord).filter(
            EtaMasterRecord.record_key == record_data['record_key'],
            EtaMasterRecord.id != record.id,
        ).first()
        if duplicate:
            flash('Another record with the same values already exists.', 'error')
            return _redirect_to_eta_master(mode='edit', page=page, per_page=per_page)

        record.sno = record_data['sno']
        record.record_key = record_data['record_key']
        record.pin_code = record_data['pin_code']
        record.pickup_station = record_data['pickup_station']
        record.state_ut = record_data['state_ut']
        record.city = record_data['city']
        record.pickup_location = record_data['pickup_location']
        record.delivery_location = record_data['delivery_location']
        record.tat_in_days = record_data['tat_in_days']
        record.zone = record_data['zone']
        record.source_filename = 'manual edit'
        record.source_row_number = None
        record.updated_at = datetime.utcnow()

        db.session.commit()
        flash('Record updated successfully.', 'success')
    except ValueError as error:
        db.session.rollback()
        flash(f'Update failed: {error}', 'error')
    except Exception as error:
        db.session.rollback()
        logger.exception('ETA master update failed')
        flash(f'Update failed: {error}', 'error')

    return _redirect_to_eta_master(mode='view', page=page, per_page=per_page)


@eta_master_bp.route('/eta-master/records/<int:record_id>/delete', methods=['POST'])
@limiter.limit('30 per minute')
def eta_master_delete_record(record_id):
    page = request.form.get('page', 1, type=int)
    per_page = request.form.get('per_page', 100, type=int)

    record = _load_eta_master_record(record_id)
    if not record:
        flash('Record not found.', 'error')
        return _redirect_to_eta_master(mode='view', page=page, per_page=per_page)

    try:
        db.session.delete(record)
        db.session.commit()
        flash('Record deleted successfully.', 'success')
    except Exception as error:
        db.session.rollback()
        logger.exception('ETA master delete failed')
        flash(f'Delete failed: {error}', 'error')

    return _redirect_to_eta_master(mode='view', page=page, per_page=per_page)


@eta_master_bp.route('/eta-master/api/count', methods=['GET'])
def eta_master_count():
    return jsonify({'count': EtaMasterRecord.query.count()})

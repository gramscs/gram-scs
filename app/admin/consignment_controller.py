"""Admin consignment management routes and helpers."""

import base64
import binascii
from datetime import date, datetime
import io
import logging
import os
import re
from uuid import uuid4

from flask import current_app, flash, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from sqlalchemy import or_
from sqlalchemy.exc import DatabaseError, OperationalError, ProgrammingError

from app import limiter
from app.admin import admin_bp
from app.admin.auth import require_admin
from app.models import Consignment, db

logger = logging.getLogger(__name__)
MAX_POD_IMAGE_BYTES = 5 * 1024 * 1024


def _get_supabase_client():
    url = os.getenv("SUPABASE_URL", "").strip()
    key = os.getenv("SUPABASE_KEY", "").strip()
    if not url or not key:
        return None
    try:
        from supabase import create_client
    except Exception:
        logger.warning("Supabase package not available; falling back to local uploads")
        return None
    try:
        return create_client(url, key)
    except Exception:
        logger.exception("Failed to create Supabase client")
        return None


def _store_pod_bytes(filename, file_bytes, content_type=None, bucket_name=None):
    supa = _get_supabase_client()
    if supa:
        bucket = bucket_name or os.getenv("SUPABASE_BUCKET", "pod-uploads")
        object_path = f"consignments/{filename}"
        payload = file_bytes
        if hasattr(payload, "read"):
            payload = payload.read()
        if isinstance(payload, bytearray):
            payload = bytes(payload)
        if not isinstance(payload, (bytes, bytearray)):
            raise TypeError("POD upload payload must be bytes-like.")
        supa.storage.from_(bucket).upload(
            object_path,
            payload,
            {"content-type": content_type or "application/octet-stream"},
        )
        return f"supabase:{bucket}/{object_path}"

    payload = file_bytes
    if hasattr(payload, "read"):
        payload = payload.read()
    if isinstance(payload, bytearray):
        payload = bytes(payload)
    if not isinstance(payload, (bytes, bytearray)):
        raise TypeError("POD upload payload must be bytes-like.")

    upload_folder = os.path.join(current_app.instance_path, "uploads")
    os.makedirs(upload_folder, exist_ok=True)
    dest_path = os.path.join(upload_folder, filename)
    with open(dest_path, "wb") as handle:
        handle.write(payload)
    return filename


def _parse_supabase_pod_value(pod_value):
    if not isinstance(pod_value, str) or not pod_value.startswith('supabase:'):
        raise ValueError('POD is not stored in Supabase.')

    _, rest = pod_value.split(':', 1)
    bucket, object_path = rest.split('/', 1)
    if not bucket or not object_path:
        raise ValueError('Invalid Supabase POD path.')

    return bucket, object_path


def _download_supabase_pod_file(pod_value):
    client = _get_supabase_client()
    if not client:
        raise RuntimeError('Supabase not configured.')

    bucket, object_path = _parse_supabase_pod_value(pod_value)
    content = client.storage.from_(bucket).download(object_path)
    if hasattr(content, 'read'):
        content = content.read()
    if isinstance(content, bytearray):
        content = bytes(content)
    if not isinstance(content, bytes):
        raise RuntimeError('Unexpected Supabase download response.')

    return content, object_path


def _download_legacy_supabase_pod_file(consignment_id, pod_value):
    """Download a legacy POD by attempting old Supabase object paths.

    Legacy records may store a bare object path or a local filename that was
    previously migrated into Supabase. This helper tries the configured bucket
    and an optional consignment-id-prefixed path.
    """
    client = _get_supabase_client()
    if not client:
        raise RuntimeError('Supabase not configured.')

    bucket = os.getenv('SUPABASE_BUCKET', 'pod-uploads')
    if not isinstance(pod_value, str) or not pod_value:
        raise ValueError('Invalid legacy POD path.')

    candidates = []
    legacy_bucket = bucket
    if pod_value.startswith('supabase:'):
        _, rest = pod_value.split(':', 1)
        try:
            legacy_bucket, legacy_object_path = rest.split('/', 1)
        except ValueError:
            legacy_bucket = bucket
            legacy_object_path = rest
        candidates.append((legacy_bucket, legacy_object_path))
    else:
        candidates.append((bucket, pod_value))
        if consignment_id is not None:
            consignment_prefix = str(consignment_id)
            if not pod_value.startswith(consignment_prefix + '/'):
                candidates.append((bucket, f"{consignment_prefix}/{pod_value}"))

    last_error = None
    for candidate_bucket, object_path in candidates:
        try:
            content = client.storage.from_(candidate_bucket).download(object_path)
            if hasattr(content, 'read'):
                content = content.read()
            if isinstance(content, bytearray):
                content = bytes(content)
            if not isinstance(content, bytes):
                raise RuntimeError('Unexpected Supabase download response.')
            return content, candidate_bucket, object_path
        except Exception as exc:
            last_error = exc

    # Last chance: if the value was a local filename under uploads, try that path.
    upload_folder = os.path.join(current_app.instance_path, 'uploads')
    try:
        legacy_path = os.path.normpath(os.path.join(upload_folder, pod_value))
        if legacy_path.startswith(os.path.abspath(upload_folder)) and os.path.exists(legacy_path):
            with open(legacy_path, 'rb') as fh:
                return fh.read(), bucket, pod_value
    except Exception:
        pass

    raise RuntimeError('Legacy POD file not found.') from last_error


def _delete_pod_file(pod_value):
    if not pod_value:
        return

    if isinstance(pod_value, str) and pod_value.startswith("supabase:"):
        client = _get_supabase_client()
        if not client:
            return
        try:
            _, rest = pod_value.split(":", 1)
            bucket, object_path = rest.split("/", 1)
            client.storage.from_(bucket).remove([object_path])
        except Exception:
            logger.exception("Failed to remove POD from Supabase")
        return

    upload_folder = os.path.join(current_app.instance_path, "uploads")
    pod_path = os.path.normpath(os.path.join(upload_folder, pod_value))
    if pod_path.startswith(os.path.abspath(upload_folder)) and os.path.exists(pod_path):
        try:
            os.remove(pod_path)
        except Exception:
            logger.exception("Failed to remove POD file from disk")


def _is_external_pod_url(value):
    if not isinstance(value, str):
        return False
    lowered = value.strip().lower()
    return lowered.startswith("http://") or lowered.startswith("https://")


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _parse_date_string(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None

    value = value.strip()
    if not value:
        return None

    try:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _serialize_consignment(consignment):
    return {
        "id": getattr(consignment, "id", None),
        "consignment_number": getattr(consignment, "consignment_number", None),
        "status": getattr(consignment, "status", None),
        "pickup_pincode": getattr(consignment, "pickup_pincode", None),
        "pickup_address": getattr(consignment, "pickup_address", None),
        "pickup_tag": getattr(consignment, "pickup_tag", None),
        "pickup_date": getattr(consignment, "pickup_date", None),
        "drop_pincode": getattr(consignment, "drop_pincode", None),
        "drop_address": getattr(consignment, "drop_address", None),
        "drop_tag": getattr(consignment, "drop_tag", None),
        "drop_date": getattr(consignment, "drop_date", None),
        "eta": getattr(consignment, "eta", None),
        "pod_image": getattr(consignment, "pod_image", None),
        "pod_file_name": getattr(consignment, "pod_file_name", None),
        "pod_file_type": getattr(consignment, "pod_file_type", None),
        "pod_file_data": getattr(consignment, "pod_file_data", None),
    }


@admin_bp.route("/admin/consignments", methods=["GET"], endpoint="consignments_panel")
@require_admin
def consignments_panel():
    try:
        total = Consignment.query.count()
        consignments = [] if total > 500 else Consignment.query.order_by(Consignment.id.asc()).limit(200).all()
        rows = [_serialize_consignment(row) for row in consignments]
        return render_template("admin/consignments.html", consignments=rows)
    except (OperationalError, DatabaseError, ProgrammingError):
        logger.exception("Database error loading admin consignments panel")
        return render_template(
            "admin/consignments.html",
            consignments=[],
            error="Unable to load data. Please try again.",
        )
    except Exception:
        logger.exception("Unexpected error in admin consignments panel")
        return render_template(
            "admin/consignments.html",
            consignments=[],
            error="An unexpected error occurred.",
        )


@admin_bp.route("/admin/consignments/list", methods=["GET"], endpoint="consignments_list_api")
@require_admin
def consignments_list_api():
    try:
        page = max(1, request.args.get("page", 1, type=int))
        per_page = max(1, min(100, request.args.get("per_page", 10, type=int)))
        search = request.args.get("search", "", type=str).strip()
        sort_by = request.args.get("sort_by", "id", type=str)
        sort_order = request.args.get("sort_order", "asc", type=str)

        allowed_sort_columns = {
            "id", "consignment_number", "status", "pickup_pincode", "drop_pincode",
            "pickup_tag", "drop_tag", "pickup_date", "drop_date",
        }
        if sort_by not in allowed_sort_columns:
            sort_by = "id"
        sort_order = "asc" if sort_order.lower() == "asc" else "desc"

        query = Consignment.query
        if search:
            pattern = f"%{search}%"
            query = query.filter(
                or_(
                    Consignment.consignment_number.ilike(pattern),
                    Consignment.status.ilike(pattern),
                    Consignment.pickup_tag.ilike(pattern),
                    Consignment.drop_tag.ilike(pattern),
                    Consignment.pickup_pincode.ilike(pattern),
                    Consignment.drop_pincode.ilike(pattern),
                    Consignment.pickup_address.ilike(pattern),
                    Consignment.drop_address.ilike(pattern),
                )
            )

        total = query.count()
        sort_column = getattr(Consignment, sort_by)
        query = query.order_by(sort_column.desc() if sort_order == "desc" else sort_column.asc())
        rows = query.offset((page - 1) * per_page).limit(per_page).all()

        pages = (total + per_page - 1) // per_page if total else 0
        return jsonify({
            "success": True,
            "rows": [_serialize_consignment(row) for row in rows],
            "page": page,
            "per_page": per_page,
            "total": total,
            "pages": pages,
            "has_prev": page > 1,
            "has_next": page < pages,
        })
    except Exception as exc:
        logger.exception("Consignment list API failed: %s", exc)
        return jsonify({
            "success": False,
            "rows": [],
            "page": 1,
            "per_page": 10,
            "total": 0,
            "pages": 0,
            "has_prev": False,
            "has_next": False,
            "error": "Unable to load consignments right now.",
        }), 500


@admin_bp.route("/admin/consignments/import-template.xlsx", methods=["GET"], endpoint="consignments_import_template_excel")
@require_admin
def consignments_import_template_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Consignments"
    sheet.append([
        "consignment_number",
        "status",
        "pickup_address",
        "pickup_pincode",
        "pickup_tag",
        "pickup_date",
        "drop_address",
        "drop_pincode",
        "drop_tag",
        "drop_date",
    ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="consignment_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route("/admin/consignments/archive", methods=["POST"], endpoint="consignments_archive")
@limiter.limit("10 per minute")
@require_admin
def consignments_archive():
    payload = request.get_json(silent=True) or {}
    before_date = str(payload.get("before_date") or "").strip()

    if not before_date:
        return jsonify({"success": False, "message": "Please provide a cutoff date."}), 400

    cutoff_date = _parse_date_string(before_date)
    if cutoff_date is None:
        return jsonify({"success": False, "message": "Cutoff date must be a valid date in YYYY-MM-DD format."}), 400

    try:
        query = Consignment.query.filter(Consignment.status.ilike("Delivered"))
        query = query.filter(Consignment.drop_date.isnot(None), Consignment.drop_date != "")

        archived_count = 0
        for consignment in query.all():
            drop_date = _parse_date_string(getattr(consignment, "drop_date", ""))
            if drop_date is None:
                continue
            if drop_date < cutoff_date:
                if getattr(consignment, "pod_image", None):
                    _delete_pod_file(consignment.pod_image)
                db.session.delete(consignment)
                archived_count += 1

        db.session.commit()
        return jsonify({"success": True, "archived_count": archived_count})
    except (OperationalError, DatabaseError):
        db.session.rollback()
        logger.exception("Database error archiving consignments")
        return jsonify({"success": False, "message": "Unable to archive consignments. Please try again."}), 500
    except Exception:
        db.session.rollback()
        logger.exception("Unexpected error archiving consignments")
        return jsonify({"success": False, "message": "An unexpected error occurred while archiving consignments."}), 500


@admin_bp.route("/admin/consignments/import", methods=["POST"], endpoint="consignments_import_excel")
@limiter.limit("10 per minute")
@require_admin
def consignments_import_excel():
    uploaded_file = request.files.get("file")
    if not uploaded_file:
        flash("Please choose an Excel file to import.", "danger")
        return redirect(url_for("admin.consignments_panel"))

    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    headers = [_normalize_header(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    existing_numbers = {
        row[0]
        for row in Consignment.query.with_entities(Consignment.consignment_number).all()
        if row and row[0]
    }

    added_count = 0
    skipped_duplicates = 0

    for row_values in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[index]: value for index, value in enumerate(row_values) if index < len(headers)}
        consignment_number = str(row_data.get("consignment_number") or "").strip()
        if not consignment_number:
            continue
        if consignment_number in existing_numbers:
            skipped_duplicates += 1
            continue

        consignment = Consignment(
            consignment_number=consignment_number,
            status=row_data.get("status") or "",
            pickup_address=row_data.get("pickup_address"),
            pickup_pincode=row_data.get("pickup_pincode"),
            pickup_tag=row_data.get("pickup_tag"),
            pickup_date=row_data.get("pickup_date"),
            drop_address=row_data.get("drop_address"),
            drop_pincode=row_data.get("drop_pincode"),
            drop_tag=row_data.get("drop_tag"),
            drop_date=row_data.get("drop_date"),
        )

        db.session.add(consignment)
        existing_numbers.add(consignment_number)
        added_count += 1

    try:
        db.session.commit()
        flash(f"Import completed. Added: {added_count}, skipped duplicates: {skipped_duplicates}.", "success")
    except Exception:
        db.session.rollback()
        logger.exception("Failed to import consignments")
        flash("Import failed.", "danger")

    return redirect(url_for("admin.consignments_panel"))


@admin_bp.route("/admin/consignments/export.xlsx", methods=["GET"], endpoint="consignments_export_excel")
@require_admin
def consignments_export_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Consignments"
    headers = [
        "consignment_number",
        "status",
        "pickup_tag",
        "drop_pincode",
        "pickup_date",
        "drop_date",
        "pickup_address",
        "drop_address",
    ]
    sheet.append(headers)

    rows = Consignment.query.order_by(Consignment.id.asc()).all()
    for consignment in rows:
        sheet.append([
            getattr(consignment, "consignment_number", None),
            getattr(consignment, "status", None),
            getattr(consignment, "pickup_tag", None),
            getattr(consignment, "drop_pincode", None),
            getattr(consignment, "pickup_date", None),
            getattr(consignment, "drop_date", None),
            getattr(consignment, "pickup_address", None),
            getattr(consignment, "drop_address", None),
        ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="consignments.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route("/admin/consignments/export.pdf", methods=["GET"], endpoint="consignments_export_pdf")
@require_admin
def consignments_export_pdf():
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    pdf.drawString(40, 550, "Consignments Export")
    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="consignments.pdf", mimetype="application/pdf")


CONSIGNMENT_SAVE_FIELDS = (
    "consignment_number",
    "status",
    "pickup_pincode",
    "pickup_address",
    "pickup_tag",
    "pickup_date",
    "drop_pincode",
    "drop_address",
    "drop_tag",
    "drop_date",
    "eta",
)


def _decode_pod_data_url(data_url):
    if not isinstance(data_url, str) or not data_url.strip():
        return None

    value = data_url.strip()
    if value.startswith("data:"):
        header, separator, encoded = value.partition(",")
        if not separator or ";base64" not in header.lower():
            raise ValueError("POD upload must be a base64 data URL.")
    else:
        encoded = value

    try:
        return base64.b64decode(encoded, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise ValueError("POD upload data is invalid.") from exc


def _pod_storage_filename(consignment_number, original_name):
    safe_original = secure_filename(original_name or "pod-upload") or "pod-upload"
    _, ext = os.path.splitext(safe_original)
    if not ext:
        ext = ".bin"
    return f"{secure_filename(consignment_number) or 'consignment'}-{uuid4().hex}{ext.lower()}"


def _apply_consignment_payload(consignment, row):
    for field in CONSIGNMENT_SAVE_FIELDS:
        value = row.get(field)
        if value is None:
            value = ""
        if isinstance(value, str):
            value = value.strip()
        setattr(consignment, field, value)


def _save_pod_upload_for_row(consignment, row, errors, row_index):
    pod_data = row.get("pod_file_data")
    if not pod_data:
        pod_image = (row.get("pod_image") or "").strip() if isinstance(row.get("pod_image"), str) else row.get("pod_image")
        if _is_external_pod_url(pod_image):
            errors.append({
                "index": row_index,
                "field": "pod_image",
                "message": "External POD URLs cannot be saved. Upload the POD file instead.",
            })
        return

    try:
        file_bytes = _decode_pod_data_url(pod_data)
    except ValueError as exc:
        errors.append({"index": row_index, "field": "pod_file_data", "message": str(exc)})
        return

    if not file_bytes:
        return

    if len(file_bytes) > MAX_POD_IMAGE_BYTES:
        errors.append({
            "index": row_index,
            "field": "pod_file_data",
            "message": "POD upload exceeds the 5 MB size limit.",
        })
        return

    if consignment.pod_image:
        _delete_pod_file(consignment.pod_image)

    filename = _pod_storage_filename(consignment.consignment_number, row.get("pod_file_name"))
    consignment.pod_image = _store_pod_bytes(filename, file_bytes, row.get("pod_file_type"))


@admin_bp.route("/admin/consignments/<int:consignment_id>/pod", methods=["POST"], endpoint="consignment_pod_upload")
@require_admin
def consignment_pod_upload(consignment_id):
    consignment = db.session.get(Consignment, consignment_id)
    if not consignment:
        return jsonify({"success": False, "message": "Consignment not found."}), 404

    uploaded_file = request.files.get("file")
    if not uploaded_file:
        return jsonify({"success": False, "message": "POD file is required."}), 400

    file_bytes = uploaded_file.read()
    if len(file_bytes) > MAX_POD_IMAGE_BYTES:
        return jsonify({"success": False, "message": "POD upload exceeds the 5 MB size limit."}), 400

    try:
        if consignment.pod_image:
            _delete_pod_file(consignment.pod_image)

        filename = _pod_storage_filename(consignment.consignment_number, uploaded_file.filename)
        consignment.pod_image = _store_pod_bytes(filename, file_bytes, uploaded_file.mimetype)
        db.session.commit()
        return jsonify({"success": True, "pod_image": consignment.pod_image})
    except Exception:
        db.session.rollback()
        logger.exception("Failed to upload POD for consignment %s", consignment_id)
        return jsonify({"success": False, "message": "Failed to upload POD."}), 500


@admin_bp.route("/admin/consignments/<int:consignment_id>/pod", methods=["GET"], endpoint="consignment_pod_download")
@require_admin
def consignment_pod_download(consignment_id):
    consignment = db.session.get(Consignment, consignment_id)
    if not consignment or not consignment.pod_image:
        return jsonify({"success": False, "message": "No POD found."}), 404

    pod_path = consignment.pod_image
    try:
        if isinstance(pod_path, str) and pod_path.startswith("supabase:"):
            content_bytes, object_path = _download_supabase_pod_file(pod_path)
            return send_file(
                io.BytesIO(content_bytes),
                as_attachment=True,
                download_name=os.path.basename(object_path) or "pod.jpg",
                mimetype="application/octet-stream",
            )

        upload_folder = os.path.join(current_app.instance_path, "uploads")
        safe_path = os.path.normpath(os.path.join(upload_folder, pod_path))
        if not safe_path.startswith(os.path.abspath(upload_folder)):
            return jsonify({"success": False, "message": "Invalid POD path."}), 400
        if not os.path.exists(safe_path):
            return jsonify({"success": False, "message": "POD file missing."}), 404
        return send_file(safe_path, as_attachment=True, download_name=os.path.basename(safe_path))
    except Exception:
        logger.exception("Failed to serve POD for consignment %s", consignment_id)
        return jsonify({"success": False, "message": "Failed to serve POD."}), 500


@admin_bp.route("/admin/consignments/<int:consignment_id>/pod", methods=["DELETE"], endpoint="consignment_pod_delete")
@require_admin
def consignment_pod_delete(consignment_id):
    consignment = db.session.get(Consignment, consignment_id)
    if not consignment:
        return jsonify({"success": False, "message": "Consignment not found."}), 404

    try:
        _delete_pod_file(consignment.pod_image)
        consignment.pod_image = None
        db.session.commit()
        return jsonify({"success": True})
    except Exception:
        db.session.rollback()
        logger.exception("Failed to delete POD for consignment %s", consignment_id)
        return jsonify({"success": False, "message": "Failed to delete POD."}), 500


def _normalize_save_payload():
    if request.is_json:
        payload = request.get_json(silent=True) or {}
    else:
        payload = request.form.to_dict(flat=True)

    if isinstance(payload.get("rows"), list):
        return payload

    # Backwards-compatible single-row payload support.
    if payload.get("consignment_number") or payload.get("id"):
        return {"rows": [payload], "deleted_ids": []}

    return payload


@admin_bp.route("/admin/consignments/save", methods=["POST"], endpoint="consignments_save")
@require_admin
def consignments_save():
    payload = _normalize_save_payload()
    rows = payload.get("rows") if isinstance(payload, dict) else None
    deleted_ids = payload.get("deleted_ids", []) if isinstance(payload, dict) else []

    if not isinstance(rows, list):
        return jsonify({"success": False, "message": "Rows payload is required."}), 400

    errors = []
    saved_count = 0
    deleted_count = 0

    try:
        for deleted_id in deleted_ids if isinstance(deleted_ids, list) else []:
            try:
                deleted_id_int = int(deleted_id)
            except (TypeError, ValueError):
                continue
            consignment = db.session.get(Consignment, deleted_id_int)
            if consignment:
                _delete_pod_file(consignment.pod_image)
                db.session.delete(consignment)
                deleted_count += 1

        for index, row in enumerate(rows):
            if not isinstance(row, dict):
                errors.append({"index": index, "field": "row", "message": "Row must be an object."})
                continue

            consignment_number = (row.get("consignment_number") or "").strip().upper()
            if not consignment_number:
                errors.append({
                    "index": index,
                    "field": "consignment_number",
                    "message": "Consignment number is required.",
                })
                continue

            row = dict(row)
            row["consignment_number"] = consignment_number

            consignment = None
            row_id = row.get("id")
            if row_id not in (None, ""):
                try:
                    consignment = db.session.get(Consignment, int(row_id))
                except (TypeError, ValueError):
                    errors.append({"index": index, "field": "id", "message": "Invalid consignment id."})
                    continue

            existing_by_number = Consignment.query.filter_by(consignment_number=consignment_number).first()
            if consignment and existing_by_number and existing_by_number.id != consignment.id:
                errors.append({
                    "index": index,
                    "field": "consignment_number",
                    "message": "Consignment number already exists.",
                })
                continue

            if not consignment:
                consignment = existing_by_number or Consignment()
                if not existing_by_number:
                    db.session.add(consignment)

            _apply_consignment_payload(consignment, row)
            _save_pod_upload_for_row(consignment, row, errors, index)
            saved_count += 1

        if errors:
            db.session.rollback()
            return jsonify({
                "success": False,
                "message": "Validation errors. Please fix highlighted rows.",
                "errors": errors,
            }), 400

        db.session.commit()
        return jsonify({
            "success": True,
            "message": "Saved.",
            "saved_count": saved_count,
            "deleted_count": deleted_count,
            "total": Consignment.query.count(),
        })
    except Exception:
        db.session.rollback()
        logger.exception("Failed to save consignments")
        return jsonify({"success": False, "message": "Failed to save consignments."}), 500

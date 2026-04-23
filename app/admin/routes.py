from flask import render_template, request, jsonify, send_file, redirect, url_for, flash
from sqlalchemy.exc import IntegrityError, OperationalError, DatabaseError
from datetime import datetime
import json
import logging
import io
import re

from openpyxl import load_workbook, Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.admin import admin_bp
from app.admin.auth import require_admin
from app.models import Consignment, Lead, db
from app.services.logistics import (
    normalize_consignment_number,
    normalize_status,
    normalize_indian_pincode,
    geocode_indian_pincode_with_retry,
    calculate_eta_breakdown_with_retry,
    get_fallback_eta,
)

logger = logging.getLogger(__name__)


@admin_bp.route("/admin/dashboard", methods=["GET"])
@require_admin
def dashboard():
    """Admin dashboard – protected landing page after login."""
    return render_template("admin/dashboard.html")


@admin_bp.route("/xk7m2p/leads", methods=["GET"])
@require_admin
def xk7m2p_leads():
    try:
        leads = Lead.query.order_by(Lead.created_at.desc(), Lead.id.desc()).all()
        rows = [
            {
                "id": lead.id,
                "name": lead.name,
                "email": lead.email,
                "phone": lead.phone,
                "subject": lead.subject,
                "message": lead.message,
                "created_at": lead.created_at.strftime("%Y-%m-%d %H:%M:%S") if lead.created_at else "",
            }
            for lead in leads
        ]
        return render_template("admin/leads.html", leads=rows)
    except (OperationalError, DatabaseError) as e:
        logger.error("Database error loading leads panel: %s", e)
        return render_template("admin/leads.html", leads=[], error="Unable to load leads right now.")
    except Exception as e:
        logger.error("Unexpected error loading leads panel: %s", e)
        return render_template("admin/leads.html", leads=[], error="An unexpected error occurred.")


def _build_eta_payload(consignment_number, pickup_pincode, drop_pincode):
    pickup_location = geocode_indian_pincode_with_retry(pickup_pincode)
    drop_location = geocode_indian_pincode_with_retry(drop_pincode)

    pickup_lat = pickup_location["lat"] if pickup_location else None
    pickup_lng = pickup_location["lng"] if pickup_location else None
    drop_lat = drop_location["lat"] if drop_location else None
    drop_lng = drop_location["lng"] if drop_location else None

    if pickup_lat is not None and pickup_lng is not None and drop_lat is not None and drop_lng is not None:
        eta_breakdown = calculate_eta_breakdown_with_retry(
            pickup_lat,
            pickup_lng,
            drop_lat,
            drop_lng,
        )
        if eta_breakdown is None:
            eta = get_fallback_eta()
            eta_breakdown = {
                "eta": eta,
                "duration_seconds": None,
                "duration_hours": None,
                "distance_km": None,
                "calculated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "route_source": "fallback",
                "formula": "ETA fallback used because route lookup failed",
            }
        else:
            eta = eta_breakdown["eta"]
    else:
        eta = get_fallback_eta()
        eta_breakdown = {
            "eta": eta,
            "duration_seconds": None,
            "duration_hours": None,
            "distance_km": None,
            "calculated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "route_source": "fallback_geocode",
            "formula": "ETA fallback used because one or more pincodes could not be geocoded",
        }

    eta_breakdown["consignment_number"] = consignment_number
    eta_breakdown["pickup_pincode"] = pickup_pincode
    eta_breakdown["drop_pincode"] = drop_pincode
    eta_breakdown["pickup_coords"] = [pickup_lat, pickup_lng]
    eta_breakdown["drop_coords"] = [drop_lat, drop_lng]
    eta_breakdown["pickup_geocode_source"] = pickup_location.get("source") if pickup_location else None
    eta_breakdown["drop_geocode_source"] = drop_location.get("source") if drop_location else None

    return {
        "pickup_lat": pickup_lat,
        "pickup_lng": pickup_lng,
        "drop_lat": drop_lat,
        "drop_lng": drop_lng,
        "eta": eta,
        "eta_debug_json": json.dumps(eta_breakdown),
    }


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


@admin_bp.route("/xk7m2p", methods=["GET"])
@require_admin
def xk7m2p():
    try:
        consignments = Consignment.query.order_by(Consignment.id.asc()).all()
        rows = [
            {
                "id": c.id,
                "consignment_number": c.consignment_number,
                "status": c.status,
                "pickup_pincode": c.pickup_pincode,
                "drop_pincode": c.drop_pincode,
                "pickup_lat": c.pickup_lat,
                "pickup_lng": c.pickup_lng,
                "drop_lat": c.drop_lat,
                "drop_lng": c.drop_lng,
                "eta": c.eta,
            }
            for c in consignments
        ]
        return render_template("admin/xk7m2p.html", consignments=rows)
    except (OperationalError, DatabaseError) as e:
        logger.error("Database error loading admin panel: %s", e)
        return render_template("admin/xk7m2p.html", consignments=[], error="Unable to load data. Please try again.")
    except Exception as e:
        logger.error("Unexpected error in admin panel: %s", e)
        return render_template("admin/xk7m2p.html", consignments=[], error="An unexpected error occurred.")


@admin_bp.route("/xk7m2p/save", methods=["POST"])
@require_admin
def xk7m2p_save():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows", [])

    if not isinstance(rows, list):
        return jsonify({"success": False, "message": "Invalid request payload."}), 400

    try:
        existing = {c.id: c for c in Consignment.query.all()}
        seen_numbers = set()
        validated_rows = []

        for row in rows:
            row_id = row.get("id")
            try:
                consignment_number = normalize_consignment_number(row.get("consignment_number"))
                status = normalize_status(row.get("status"))
                pickup_pincode = normalize_indian_pincode(row.get("pickup_pincode"), "pickup_pincode")
                drop_pincode = normalize_indian_pincode(row.get("drop_pincode"), "drop_pincode")
            except ValueError as error:
                return jsonify({"success": False, "message": str(error)}), 400

            if consignment_number in seen_numbers:
                return jsonify({
                    "success": False,
                    "message": f"Duplicate consignment number in sheet: {consignment_number}"
                }), 400
            seen_numbers.add(consignment_number)

            eta_payload = _build_eta_payload(consignment_number, pickup_pincode, drop_pincode)

            if row_id:
                try:
                    row_id = int(row_id)
                except (TypeError, ValueError):
                    return jsonify({"success": False, "message": f"Invalid row id: {row_id}"}), 400

                if row_id not in existing:
                    return jsonify({"success": False, "message": f"Row id {row_id} not found."}), 400
            else:
                row_id = None

            validated_rows.append({
                "id": row_id,
                "consignment_number": consignment_number,
                "status": status,
                "pickup_pincode": pickup_pincode,
                "drop_pincode": drop_pincode,
                "pickup_lat": eta_payload["pickup_lat"],
                "pickup_lng": eta_payload["pickup_lng"],
                "drop_lat": eta_payload["drop_lat"],
                "drop_lng": eta_payload["drop_lng"],
                "eta": eta_payload["eta"],
                "eta_debug_json": eta_payload["eta_debug_json"],
            })

        for row in validated_rows:
            if row["id"]:
                consignment = existing[row["id"]]
            else:
                consignment = Consignment()
                db.session.add(consignment)

            consignment.consignment_number = row["consignment_number"]
            consignment.status = row["status"]
            consignment.pickup_pincode = row["pickup_pincode"]
            consignment.drop_pincode = row["drop_pincode"]
            consignment.pickup_lat = row["pickup_lat"]
            consignment.pickup_lng = row["pickup_lng"]
            consignment.drop_lat = row["drop_lat"]
            consignment.drop_lng = row["drop_lng"]
            consignment.eta = row["eta"]
            consignment.eta_debug_json = row["eta_debug_json"]

        db.session.commit()
        return jsonify({"success": True, "message": "Sheet saved successfully."})

    except IntegrityError as e:
        db.session.rollback()
        logger.error("Integrity error in admin save: %s", e)
        return jsonify({"success": False, "message": "Duplicate consignment number already exists."}), 400
    except (OperationalError, DatabaseError) as e:
        db.session.rollback()
        logger.error("Database error in admin save: %s", e)
        return jsonify({"success": False, "message": "Database connection error. Please try again."}), 500
    except ValueError as e:
        db.session.rollback()
        logger.error("Validation error in admin save: %s", e)
        return jsonify({"success": False, "message": str(e)}), 400
    except Exception as e:
        db.session.rollback()
        logger.error("Unexpected error in admin save: %s", e)
        return jsonify({"success": False, "message": "An unexpected error occurred. Please try again."}), 500


@admin_bp.route("/xk7m2p/import", methods=["POST"])
@require_admin
def xk7m2p_import_excel():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        flash("Please choose an Excel file (.xlsx).", "danger")
        return redirect(url_for("admin.xk7m2p"))

    filename = upload.filename.lower()
    if not filename.endswith(".xlsx"):
        flash("Only .xlsx files are supported.", "danger")
        return redirect(url_for("admin.xk7m2p"))

    try:
        workbook = load_workbook(upload, data_only=True)
        sheet = workbook.active

        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            flash("Excel file is empty.", "danger")
            return redirect(url_for("admin.xk7m2p"))

        normalized_headers = [_normalize_header(cell) for cell in header_cells]
        header_index = {name: idx for idx, name in enumerate(normalized_headers) if name}

        consignment_idx = header_index.get("consignment_number")
        status_idx = header_index.get("status")
        pickup_idx = header_index.get("pickup_pincode")
        drop_idx = header_index.get("drop_pincode")

        if None in (consignment_idx, status_idx, pickup_idx, drop_idx):
            flash("Required headers: consignment_number, status, pickup_pincode, drop_pincode", "danger")
            return redirect(url_for("admin.xk7m2p"))

        existing_numbers = {c.consignment_number for c in Consignment.query.with_entities(Consignment.consignment_number).all()}
        file_seen = set()
        added_count = 0
        skipped_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(value is None or str(value).strip() == "" for value in row):
                continue

            consignment_number = normalize_consignment_number(row[consignment_idx])
            status = normalize_status(row[status_idx])
            pickup_pincode = normalize_indian_pincode(row[pickup_idx], "pickup_pincode")
            drop_pincode = normalize_indian_pincode(row[drop_idx], "drop_pincode")

            if consignment_number in existing_numbers or consignment_number in file_seen:
                skipped_count += 1
                continue

            eta_payload = _build_eta_payload(consignment_number, pickup_pincode, drop_pincode)

            consignment = Consignment(
                consignment_number=consignment_number,
                status=status,
                pickup_pincode=pickup_pincode,
                drop_pincode=drop_pincode,
                pickup_lat=eta_payload["pickup_lat"],
                pickup_lng=eta_payload["pickup_lng"],
                drop_lat=eta_payload["drop_lat"],
                drop_lng=eta_payload["drop_lng"],
                eta=eta_payload["eta"],
                eta_debug_json=eta_payload["eta_debug_json"],
            )

            db.session.add(consignment)
            file_seen.add(consignment_number)
            existing_numbers.add(consignment_number)
            added_count += 1

        db.session.commit()
        flash(f"Import completed. Added: {added_count}, skipped duplicates: {skipped_count}.", "success")
        return redirect(url_for("admin.xk7m2p"))
    except ValueError as e:
        db.session.rollback()
        flash(str(e), "danger")
        return redirect(url_for("admin.xk7m2p"))
    except Exception as e:
        db.session.rollback()
        logger.error("Unexpected error in Excel import: %s", e)
        flash("Failed to import Excel file.", "danger")
        return redirect(url_for("admin.xk7m2p"))


@admin_bp.route("/xk7m2p/export.xlsx", methods=["GET"])
@require_admin
def xk7m2p_export_excel():
    try:
        rows = Consignment.query.order_by(Consignment.id.asc()).all()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Internal Consignments"

        sheet.append([
            "id",
            "consignment_number",
            "status",
            "pickup_pincode",
            "drop_pincode",
            "pickup_lat",
            "pickup_lng",
            "drop_lat",
            "drop_lng",
            "eta",
        ])

        for row in rows:
            sheet.append([
                row.id,
                row.consignment_number,
                row.status,
                row.pickup_pincode,
                row.drop_pincode,
                row.pickup_lat,
                row.pickup_lng,
                row.drop_lat,
                row.drop_lng,
                row.eta,
            ])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="internal_consignments.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logger.error("Excel export failed: %s", e)
        return jsonify({"success": False, "message": "Failed to export Excel."}), 500


@admin_bp.route("/xk7m2p/export.pdf", methods=["GET"])
@require_admin
def xk7m2p_export_pdf():
    try:
        rows = Consignment.query.order_by(Consignment.id.asc()).all()

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()

        table_data = [["ID", "Consignment #", "Status", "Pickup", "Drop", "ETA"]]
        for row in rows:
            table_data.append([
                str(row.id),
                row.consignment_number or "",
                row.status or "",
                row.pickup_pincode or "",
                row.drop_pincode or "",
                row.eta or "",
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E9ECEF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#6C757D")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        content = [
            Paragraph("Internal Consignment Sheet", styles["Heading2"]),
            Spacer(1, 8),
            table,
        ]

        doc.build(content)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="internal_consignments.pdf",
            mimetype="application/pdf",
        )
    except Exception as e:
        logger.error("PDF export failed: %s", e)
        return jsonify({"success": False, "message": "Failed to export PDF."}), 500


@admin_bp.route("/xk7m2p/import-template.xlsx", methods=["GET"])
@require_admin
def xk7m2p_import_template_excel():
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Import Template"

        sheet.append([
            "consignment_number",
            "status",
            "pickup_pincode",
            "drop_pincode",
        ])

        sheet.append([
            "CN001",
            "In Transit",
            "110017",
            "400001",
        ])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="internal_consignments_import_template.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logger.error("Template export failed: %s", e)
        return jsonify({"success": False, "message": "Failed to generate import template."}), 500

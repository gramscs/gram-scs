import unittest
from io import BytesIO
from openpyxl import Workbook

from app import create_app, db
from app.eta_master.models import EtaMasterRecord


class EtaMasterHeaderDebugTests(unittest.TestCase):
    """Test header mapping with various Excel file formats."""

    def setUp(self):
        self.app = create_app()
        self.app.config['TESTING'] = True
        self.client = self.app.test_client()

        with self.app.app_context():
            db.session.query(EtaMasterRecord).delete()
            db.session.commit()

    def tearDown(self):
        with self.app.app_context():
            db.session.query(EtaMasterRecord).delete()
            db.session.commit()

    def test_upload_with_standard_headers(self):
        """Test upload with standard headers (SNO, PIN CODE, etc.)"""
        wb = Workbook()
        ws = wb.active
        ws.append(['SNO', 'PIN CODE', 'PICK UP STATION', 'STATE/UT', 'CITY', 'PICK UP LOCATION', 'DELIVERY LOCATION', 'TAT IN DAYS', 'ZONE'])
        ws.append([1, '560001', 'Bengaluru Hub', 'Karnataka', 'Bangalore', 'Station', 'Zone', 1.5, 'South'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = self.client.post(
            '/eta-master',
            data={'file': (output, 'test.xlsx')},
            content_type='multipart/form-data',
        )

        self.assertIn(b'Inserted', response.data)

    def test_upload_with_lowercase_headers(self):
        """Test upload with lowercase headers."""
        wb = Workbook()
        ws = wb.active
        ws.append(['sno', 'pin code', 'pick up station', 'state/ut', 'city', 'pick up location', 'delivery location', 'tat in days', 'zone'])
        ws.append([1, '560001', 'Bengaluru Hub', 'Karnataka', 'Bangalore', 'Station', 'Zone', 1.5, 'South'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = self.client.post(
            '/eta-master',
            data={'file': (output, 'test.xlsx')},
            content_type='multipart/form-data',
        )

        self.assertIn(b'Inserted', response.data)

    def test_upload_with_mixed_case_headers(self):
        """Test upload with mixed case headers."""
        wb = Workbook()
        ws = wb.active
        ws.append(['Sno', 'Pin Code', 'Pick Up Station', 'State/UT', 'City', 'Pick Up Location', 'Delivery Location', 'TAT in Days', 'Zone'])
        ws.append([1, '560001', 'Bengaluru Hub', 'Karnataka', 'Bangalore', 'Station', 'Zone', 1.5, 'South'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = self.client.post(
            '/eta-master',
            data={'file': (output, 'test.xlsx')},
            content_type='multipart/form-data',
        )

        self.assertIn(b'Inserted', response.data)

    def test_upload_with_extra_spaces(self):
        """Test upload with extra spaces in headers."""
        wb = Workbook()
        ws = wb.active
        ws.append(['  SNO  ', '  PIN CODE  ', '  PICK UP STATION  ', '  STATE/UT  ', '  CITY  ', '  PICK UP LOCATION  ', '  DELIVERY LOCATION  ', '  TAT IN DAYS  ', '  ZONE  '])
        ws.append([1, '560001', 'Bengaluru Hub', 'Karnataka', 'Bangalore', 'Station', 'Zone', 1.5, 'South'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = self.client.post(
            '/eta-master',
            data={'file': (output, 'test.xlsx')},
            content_type='multipart/form-data',
        )

        self.assertIn(b'Inserted', response.data)

    def test_upload_with_abbreviated_headers(self):
        """Test upload with abbreviated headers (TAT, PIN, etc.)"""
        wb = Workbook()
        ws = wb.active
        ws.append(['SNO', 'PIN', 'STATION', 'STATE', 'CITY', 'PICKUP', 'DELIVERY', 'TAT', 'ZONE'])
        ws.append([1, '560001', 'Bengaluru Hub', 'Karnataka', 'Bangalore', 'Station', 'Zone', 1.5, 'South'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = self.client.post(
            '/eta-master',
            data={'file': (output, 'test.xlsx')},
            content_type='multipart/form-data',
        )

        # Should work with the expanded aliases
        self.assertIn(b'Inserted', response.data)

    def test_upload_with_hyphens_in_headers(self):
        """Test upload with hyphens in headers (Pick-up, etc.)"""
        wb = Workbook()
        ws = wb.active
        ws.append(['SNO', 'PIN CODE', 'PICK-UP STATION', 'STATE/UT', 'CITY', 'PICK-UP LOCATION', 'DELIVERY LOCATION', 'TAT IN DAYS', 'ZONE'])
        ws.append([1, '560001', 'Bengaluru Hub', 'Karnataka', 'Bangalore', 'Station', 'Zone', 1.5, 'South'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = self.client.post(
            '/eta-master',
            data={'file': (output, 'test.xlsx')},
            content_type='multipart/form-data',
        )

        # Should work since spaces are normalized
        self.assertIn(b'Inserted', response.data)


if __name__ == '__main__':
    unittest.main()

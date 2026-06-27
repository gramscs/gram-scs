import io
import os
import unittest
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from app import create_app
from openpyxl import Workbook, load_workbook


class AdminConsignmentImportExportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        os.environ.setdefault("SECRET_KEY", "test-secret-key")
        os.environ.setdefault("ADMIN_USERNAME", "admin")
        os.environ.setdefault("ADMIN_PASSWORD_HASH", "scrypt:32768:8:1$yFUNQ6eCe1ScMEcQ$d94441786edd350236b9340455e3302df2cbb8cf12ba94311abf8d2f03f55325f913b75a20efc1c7a7a8ffaa0357c3b9e0f246dea70c4ea368f0346072f03f55325f913b")

        with patch("app._require_database_uri", return_value="postgresql://user:pass@localhost/testdb"), patch("app._should_auto_create_tables", return_value=False):
            cls.app = create_app()
        cls.app.config.update(TESTING=True)
        cls.client = cls.app.test_client()

    def test_import_template_contains_modal_columns(self):
        with patch("app.admin.auth.is_admin_authenticated", return_value=True):
            response = self.client.get("/admin/consignments/import-template.xlsx")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(io.BytesIO(response.data), data_only=True)
        sheet = workbook.active
        headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

        self.assertEqual(headers, [
            "consignment_number",
            "status",
            "pickup_address",
            "pickup_pincode",
            "pickup_tag",
            "pickup_date",
            "drop_address",
            "drop_pincode",
            "drop_tag",
            "drop_date",
        ])

    def test_import_excel_accepts_modal_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "consignment_number",
            "status",
            "pickup_address",
            "pickup_pincode",
            "pickup_tag",
            "pickup_date",
            "drop_address",
            "drop_pincode",
            "drop_tag",
            "drop_date",
        ])
        sheet.append([
            "CN1234",
            "In Transit",
            "123 Main Street, New Delhi",
            "110017",
            "PICKUP-001",
            "2026-05-10",
            "456 Marine Drive, Mumbai",
            "400001",
            "DROP-001",
            "2026-05-12",
        ])

        file_stream = io.BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        fake_query = MagicMock()
        fake_query.with_entities.return_value.all.return_value = []

        with patch("app.admin.auth.is_admin_authenticated", return_value=True), patch("app.admin.consignment_controller.Consignment") as FakeConsignment, patch("app.admin.consignment_controller.db") as fake_db:
            FakeConsignment.query = fake_query
            fake_db.session.add.return_value = None
            fake_db.session.commit.return_value = None
            fake_db.session.rollback.return_value = None

            response = self.client.post(
                "/admin/consignments/import",
                data={"file": (file_stream, "import.xlsx")},
                content_type="multipart/form-data",
                follow_redirects=True,
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Import completed. Added: 1, skipped duplicates: 0.", response.data)
        FakeConsignment.assert_called_once()

    def test_export_excel_has_screen_columns_plus_address(self):
        fake_row = SimpleNamespace(
            consignment_number="CN1234",
            status="Delivered",
            pickup_tag="PICKUP-001",
            drop_pincode="400001",
            pickup_date="2026-05-10",
            drop_date="2026-05-12",
            pickup_address="123 Main Street, New Delhi",
            drop_address="456 Marine Drive, Mumbai",
        )

        fake_query = MagicMock()
        fake_query.order_by.return_value.all.return_value = [fake_row]

        with patch("app.admin.auth.is_admin_authenticated", return_value=True), patch("app.admin.consignment_controller.Consignment") as FakeConsignment:
            FakeConsignment.query = fake_query
            response = self.client.get("/admin/consignments/export.xlsx")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(io.BytesIO(response.data), data_only=True)
        sheet = workbook.active
        headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        self.assertEqual(headers, [
            "consignment_number",
            "status",
            "pickup_tag",
            "drop_pincode",
            "pickup_date",
            "drop_date",
            "pickup_address",
            "drop_address",
        ])

        row_values = [cell for cell in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
        self.assertEqual(row_values, [
            "CN1234",
            "Delivered",
            "PICKUP-001",
            "400001",
            "2026-05-10",
            "2026-05-12",
            "123 Main Street, New Delhi",
            "456 Marine Drive, Mumbai",
        ])


if __name__ == "__main__":
    unittest.main()

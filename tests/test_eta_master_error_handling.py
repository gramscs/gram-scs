import unittest
from io import BytesIO
from openpyxl import Workbook

from app import create_app, db
from app.eta_master.models import EtaMasterRecord


class EtaMasterErrorHandlingTests(unittest.TestCase):
    """Test error handling for various wrong header formats."""

    def setUp(self):
        self.app = create_app()
        self.app.config['TESTING'] = True
        self.client = self.app.test_client()

        with self.app.app_context():
            db.session.query(EtaMasterRecord).delete()
            db.session.commit()

    def tearDown(self):
        with self.app.app_context():
            db.session.query(EtaMasterRecord).delete()
            db.session.commit()

    def test_upload_with_completely_wrong_headers(self):
        """Test that uploading a file with completely wrong headers shows proper error."""
        wb = Workbook()
        ws = wb.active
        # Wrong headers that don't match any aliases
        ws.append(['ID', 'POSTCODE', 'NAME', 'REGION', 'LOCATION', 'WAREHOUSE', 'DESTINATION', 'DAYS', 'AREA'])
        ws.append([1, '560001', 'Bengaluru', 'Karnataka', 'Main', 'Hub', 'Zone', 1.5, 'South'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = self.client.post(
            '/eta-master',
            data={'file': (output, 'wrong_headers.xlsx')},
            content_type='multipart/form-data',
        )

        # Should show error
        self.assertIn(b'Missing required columns', response.data)
        print("✓ Error handling test passed - proper error message shown")

    def test_upload_with_missing_one_column(self):
        """Test that uploading with one missing column shows error."""
        wb = Workbook()
        ws = wb.active
        # Missing ZONE column
        ws.append(['SNO', 'PIN CODE', 'PICK UP STATION', 'STATE/UT', 'CITY', 'PICK UP LOCATION', 'DELIVERY LOCATION', 'TAT IN DAYS'])
        ws.append([1, '560001', 'Bengaluru Hub', 'Karnataka', 'Bangalore', 'Station', 'Zone', 1.5])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = self.client.post(
            '/eta-master',
            data={'file': (output, 'missing_zone.xlsx')},
            content_type='multipart/form-data',
        )

        # Should show error mentioning zone
        self.assertIn(b'zone', response.data)
        print("✓ Missing column detection works - error mentions 'zone'")

    def test_upload_with_all_correct_headers_lowercase(self):
        """Test successful upload with all correct headers in lowercase."""
        wb = Workbook()
        ws = wb.active
        ws.append(['sno', 'pin code', 'pick up station', 'state/ut', 'city', 'pick up location', 'delivery location', 'tat in days', 'zone'])
        ws.append([1, '560001', 'Bengaluru Hub', 'Karnataka', 'Bangalore', 'Station', 'Zone', 1.5, 'South'])
        ws.append([2, '400001', 'Mumbai Hub', 'Maharashtra', 'Mumbai', 'Port', 'Warehouse', 2.0, 'West'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = self.client.post(
            '/eta-master',
            data={'file': (output, 'correct_lowercase.xlsx')},
            content_type='multipart/form-data',
        )

        # Should succeed
        self.assertIn(b'Inserted 2', response.data)
        print("✓ Lowercase headers work - 2 records inserted")

    def test_error_message_is_user_friendly(self):
        """Test that error messages clearly show which columns are missing."""
        wb = Workbook()
        ws = wb.active
        # Missing multiple columns
        ws.append(['PIN', 'CITY'])
        ws.append(['560001', 'Bangalore'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = self.client.post(
            '/eta-master',
            data={'file': (output, 'minimal.xlsx')},
            content_type='multipart/form-data',
        )

        # Check that error message is clear
        self.assertIn(b'Missing required columns:', response.data)
        
        # Should list the missing columns
        missing_columns = [b'pin_code', b'pickup_station', b'state_ut', b'city', 
                          b'pickup_location', b'delivery_location', b'tat_in_days', b'zone']
        found_any = False
        for col in missing_columns:
            if col in response.data:
                found_any = True
                break
        
        self.assertTrue(found_any)
        print("✓ Error message is user-friendly and lists missing columns")


if __name__ == '__main__':
    unittest.main()
